import openpyxl
import re
from django.db.models import Count, Q
from django.http import HttpResponse
from django.utils.decorators import method_decorator
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
from openpyxl.drawing.image import Image
from openpyxl.writer.excel import save_virtual_workbook
from rest_framework import mixins
from rest_framework import permissions
from rest_framework.exceptions import APIException
from rest_framework.parsers import MultiPartParser
from rest_framework.viewsets import GenericViewSet

from bothub import utils, settings
from bothub.api.v2.metadata import Metadata
from bothub.api.v2.mixins import MultipleFieldLookupMixin
from bothub.api.v2.translation.filters import TranslationsFilter
from bothub.api.v2.translation.permissions import (
    RepositoryTranslatedExamplePermission,
    RepositoryTranslatedExampleExporterPermission,
)
from bothub.api.v2.translation.serializers import (
    RepositoryTranslatedExampleSerializer,
    RepositoryTranslatedExporterSerializer,
    RepositoryTranslatedImportSerializer,
)
from bothub.common.models import RepositoryExample
from bothub.common.models import RepositoryExampleEntity
from bothub.common.models import (
    RepositoryTranslatedExample,
    RepositoryVersion,
    RepositoryTranslatedExampleEntity,
)


class RepositoryTranslatedExampleViewSet(
    mixins.CreateModelMixin,
    mixins.ListModelMixin,
    mixins.RetrieveModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin,
    GenericViewSet,
):
    """
    Manager example translation.

    list:
    Get example translation data.

    retrieve:
    Get example translation data.

    update:
    Update example translation.

    partial_update:
    Update, partially, example translation.

    delete:
    Delete example translation.
    """

    queryset = RepositoryTranslatedExample.objects
    serializer_class = RepositoryTranslatedExampleSerializer
    permission_classes = [
        permissions.IsAuthenticatedOrReadOnly,
        RepositoryTranslatedExamplePermission,
    ]

    def create(self, request, *args, **kwargs):
        self.permission_classes = [permissions.IsAuthenticated]
        return super().create(request, *args, **kwargs)

    def list(self, request, *args, **kwargs):
        self.queryset = RepositoryTranslatedExample.objects.all()
        self.filter_class = TranslationsFilter
        return super().list(request, *args, **kwargs)


@method_decorator(
    name="retrieve",
    decorator=swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter(
                "with_translation",
                openapi.IN_QUERY,
                description="Boolean that allows you to download or not all translations",
                type=openapi.TYPE_BOOLEAN,
                default=True,
            ),
            openapi.Parameter(
                "of_the_language",
                openapi.IN_QUERY,
                description="Choose the language to be translated",
                type=openapi.TYPE_STRING,
                required=True,
            ),
            openapi.Parameter(
                "for_the_language",
                openapi.IN_QUERY,
                description="Which language will be translated",
                type=openapi.TYPE_STRING,
                required=True,
            ),
        ]
    ),
)
class RepositoryTranslatedExporterViewSet(
    MultipleFieldLookupMixin,
    mixins.RetrieveModelMixin,
    mixins.UpdateModelMixin,
    GenericViewSet,
):

    queryset = RepositoryVersion.objects
    lookup_field = "repository__uuid"
    lookup_fields = ["repository__uuid", "pk"]
    serializer_class = RepositoryTranslatedExporterSerializer
    permission_classes = [
        permissions.IsAuthenticated,
        RepositoryTranslatedExampleExporterPermission,
    ]
    parser_classes = (MultiPartParser,)
    metadata_class = Metadata

    def retrieve(self, request, *args, **kwargs):  # pragma: no cover
        repository_version = self.get_object()

        serializer = RepositoryTranslatedImportSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)

        with_translation = serializer.data.get("with_translation")
        of_the_language = serializer.data.get("of_the_language")
        for_the_language = serializer.data.get("for_the_language")

        queryset = RepositoryExample.objects.filter(
            repository_version_language__repository_version=repository_version
        )

        examples = repository_version.repository.examples(
            queryset=queryset, version_default=repository_version.is_default
        ).annotate(
            translation_count=Count(
                "translations", filter=Q(translations__language=for_the_language)
            )
        )

        if with_translation:
            examples = examples.filter(
                repository_version_language__language=of_the_language
            )
        else:
            examples = examples.filter(
                repository_version_language__language=of_the_language,
                translation_count=0,
            )

        examples = examples.order_by("created_at")

        workbook = openpyxl.load_workbook(
            f"{settings.STATIC_ROOT}/bothub/exporter/example.xlsx"
        )
        worksheet = workbook.get_sheet_by_name("Translate")

        img = Image(f"{settings.STATIC_ROOT}/bothub/exporter/bothub.png")
        img.width = 210.21
        img.height = 60.69

        worksheet.add_image(img, "B1")

        entities_list = []

        for count, example in enumerate(examples, start=14):
            worksheet.insert_rows(count)
            worksheet.cell(row=count, column=2, value=str(example.pk))
            worksheet.cell(
                row=count,
                column=3,
                value=str(example.repository_version_language.repository_version.pk),
            )
            worksheet.cell(
                row=count,
                column=4,
                value=str(example.repository_version_language.language),
            )

            text = example.text
            entities = RepositoryExampleEntity.objects.filter(
                repository_example=example
            ).order_by("start")
            count_entity = 0
            for entity in entities:
                if entity.entity.value not in entities_list:
                    entities_list.append(entity.entity.value)
                text = utils.format_entity(
                    text=text,
                    entity=entity.entity.value,
                    start=entity.start + count_entity,
                    end=entity.end + count_entity,
                )
                count_entity += len(entity.entity.value) + 4
            worksheet.cell(row=count, column=5, value=str(text))

            translated = RepositoryTranslatedExample.objects.filter(
                original_example=example.pk, language=for_the_language
            )
            if translated:
                translated = translated.first()
                text_translated = translated.text
                entities_translate = RepositoryTranslatedExampleEntity.objects.filter(
                    repository_translated_example=translated
                ).order_by("start")
                count_entity = 0
                for entity in entities_translate:
                    text_translated = utils.format_entity(
                        text=text_translated,
                        entity=entity.entity.value,
                        start=entity.start + count_entity,
                        end=entity.end + count_entity,
                    )
                    count_entity += len(entity.entity.value) + 4
                worksheet.cell(row=count, column=6, value=str(text_translated))

        for count, entity in enumerate(entities_list, start=4):
            worksheet.insert_rows(count)
            worksheet.cell(row=count, column=2, value=str(entity))

        response = HttpResponse(
            content=save_virtual_workbook(workbook),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename=bothub.xlsx"
        return response

    def update(self, request, *args, **kwargs):  # pragma: no cover
        serializer = RepositoryTranslatedExporterSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        for_language = serializer.data.get("language", None)

        workbook = openpyxl.load_workbook(filename=request.data.get("file"))
        worksheet = workbook.get_sheet_by_name("Translate")
        columns = ["ID", "Repository Version", "Language"]

        examples_success = []

        find = False
        for count, row in enumerate(worksheet.iter_rows(), start=1):
            if (
                row[1].value in columns
                and row[2].value in columns
                and row[3].value in columns
            ):
                find = True
                continue
            if find:
                example_id = int(re.sub("[^0-9]", "", row[1].value))
                repository_version = int(row[2].value)
                text_translated = row[5].value

                if not int(kwargs.get("pk")) == repository_version:
                    raise APIException(  # pragma: no cover
                        {
                            "detail": "Import version is different from the selected version"
                        },
                        code=400,
                    )

                if text_translated:
                    example = RepositoryExample.objects.filter(
                        pk=example_id,
                        repository_version_language__repository_version=repository_version,
                        repository_version_language__repository_version__repository=kwargs.get(
                            "repository__uuid"
                        ),
                    )
                    if example.count() == 0:
                        worksheet.cell(
                            row=count, column=7, value="Sentence does not exist"
                        )
                        continue

                    example = example.first()

                    entity_validation = False

                    for entity in utils.find_entities_in_example(text_translated):
                        original_text_count_entity = RepositoryExampleEntity.objects.filter(
                            repository_example=example,
                            entity__repository_version=kwargs.get("pk"),
                            entity__value=entity.get("entity"),
                        ).count()

                        if original_text_count_entity == 0:
                            entity_validation = True
                            worksheet.cell(
                                row=count, column=7, value="Entities must match"
                            )
                            break

                    if entity_validation:
                        continue

                    translated_examples = RepositoryTranslatedExample.objects.filter(
                        original_example=example, language=for_language
                    )

                    if translated_examples.count() > 0:
                        translated_examples.delete()

                    version_language = example.repository_version_language.repository_version.repository.current_version(
                        language=for_language,
                        is_default=example.repository_version_language.repository_version.is_default,
                    )

                    translated = RepositoryTranslatedExample.objects.create(
                        repository_version_language=version_language,
                        original_example=example,
                        language=for_language,
                        text=utils.get_without_entity(text_translated),
                        clone_repository=True,
                    )

                    for translated_entity in utils.find_entities_in_example(
                        text_translated
                    ):
                        RepositoryTranslatedExampleEntity.objects.create(
                            repository_translated_example=translated,
                            start=translated_entity["start"],
                            end=translated_entity["end"],
                            entity=translated_entity["entity"],
                        )

                examples_success.append(count)

        for r in reversed(examples_success):
            worksheet.delete_rows(r)

        response = HttpResponse(
            content=save_virtual_workbook(workbook),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename=bothub.xlsx"
        return response
