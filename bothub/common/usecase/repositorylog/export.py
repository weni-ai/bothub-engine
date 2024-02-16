import io

from rest_framework import status
from rest_framework.response import Response

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

from .storage import s3FileDatabase


class ExportRepositoryLogUseCase:

    def _create_xlsx_workbook(
            self,
            repository_logs
    ) -> Workbook:

        wb = Workbook()
        ws = wb.active

        ws['A1'] = 'Text'
        ws['B1'] = 'Created At'
        ws['C1'] = 'Intent'
        ws['D1'] = 'Confidence'
        ws['E1'] = 'Entities'
        ws['F1'] = 'Entities List'

        if repository_logs is None:
            return wb

        row = 2
        for repository_log in repository_logs:
            entities = repository_log.nlp_log.entities.to_dict()
            entities_list = []
            for entity_type, entities in entities.items():
                for entity in entities:
                    entities_list.append(f"{entity['entity']}:{entity['value']}")
            entities_str = ', '.join(entities_list)

            ws['A{}'.format(row)] = repository_log.nlp_log.text
            ws['B{}'.format(row)] = repository_log.created_at
            ws['C{}'.format(row)] = repository_log.nlp_log.intent.name
            ws['D{}'.format(row)] = repository_log.nlp_log.intent.confidence
            ws['F{}'.format(row)] = entities_str
            row += 1

        return wb

    def xlsx_storage(
            self,
            xlsx_file
    ) -> str:
        file_database = s3FileDatabase()
        xlsx_file_object = io.BytesIO(xlsx_file)
        created_file = file_database.add_file(xlsx_file_object)
        temp_url = file_database.create_presigned_url(created_file.file_name)
        return temp_url

    def create_xlsx_response(
            self,
            repository_logs
    ) -> Response:

        wb = self._create_xlsx_workbook(repository_logs)
        temp_url = self.xlsx_storage(save_virtual_workbook(wb))
        response = Response(
            data={"file": temp_url},
            status=status.HTTP_200_OK
        )

        return response
